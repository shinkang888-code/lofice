#!/usr/bin/env python3
"""excel-convert API — Phase 3 XLSX 분석·변환"""

from __future__ import annotations

import base64
import io
import os
from typing import Any

import uvicorn
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware

app = FastAPI(title="lofice excel-convert API", version="1.0.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/health")
def health() -> dict[str, Any]:
    return {"status": "ok", "features": ["analyze", "to-csv"]}


@app.post("/analyze")
async def analyze(file: UploadFile = File(...)) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=503, detail="openpyxl not installed")

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="empty file")
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        preview = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 10:
                break
            preview.append([str(c) if c is not None else None for c in row])
        sheets.append({
            "name": name,
            "rows": ws.max_row,
            "cols": ws.max_column,
            "preview": preview,
        })
    wb.close()
    return {
        "file_name": file.filename,
        "sheet_count": len(sheets),
        "sheets": sheets,
        "compatibility": {"cells": "full", "charts": "none", "pivot": "none", "vba": "none"},
    }


@app.post("/to-csv")
async def to_csv(file: UploadFile = File(...)) -> dict[str, str]:
    try:
        import openpyxl
        import csv
    except ImportError:
        raise HTTPException(status_code=503, detail="openpyxl not installed")

    data = await file.read()
    if not data:
        raise HTTPException(status_code=400, detail="empty file")
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    buf = io.StringIO()
    writer = csv.writer(buf)
    for row in ws.iter_rows(values_only=True):
        writer.writerow(["" if c is None else c for c in row])
    wb.close()
    csv_bytes = buf.getvalue().encode("utf-8-sig")
    base = (file.filename or "workbook").rsplit(".", 1)[0]
    return {
        "file_name": f"{base}.csv",
        "data_base64": base64.b64encode(csv_bytes).decode(),
        "format": "csv",
    }


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=int(os.environ.get("PORT", "8400")))
